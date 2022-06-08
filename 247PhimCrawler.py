
from VideoDownloader import m3u8DownloadIDM
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
import time
import sys
import json
from unidecode import unidecode
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote import webelement
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from browsermobproxy import Server
from ProxyManagerModule import ProxyManager
import pprint
from urllib.parse import unquote
import re
from bs4 import BeautifulSoup
import openpyxl
proxy = ProxyManager()
server = proxy.start_server()
client = proxy.start_client()
time.sleep(4)
driver = ""
m3u8FileParReg = re.compile(r".m3u8")
class PhimmoiCrawler : 
    
    def __init__(self,url) -> None:
        global driver
        global client
        global server
        options = Options()
        options.add_extension("./bin/IDM-Integration-Module_v6.38.19.crx")
        options.add_argument(f"--proxy-server={client.proxy}")
        options.add_argument('--ignore-ssl-errors=yes')
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--disable-gpu')
        options.add_argument("--proxy-bypass-list=*247phim*")
        driver = webdriver.Chrome(executable_path=r"./bin/chromedriver", chrome_options=options,desired_capabilities=options.to_capabilities())
        driver.get(url)
        time.sleep(4)
    
   
    def interactSiteToAjaxCome(self,count_times):
        global driver
        for i in range(0,count_times) :
           driver.find_element_by_tag_name("body").send_keys(Keys.PAGE_DOWN)
           time.sleep(1)
    def GoIntoFilm(self) : 
        global driver
        global client
        global server
        index = sys.argv[1] or input("Type a index to start from")
        for page in range(2, 45) : 
            
            films_item = driver.find_elements_by_css_selector(".list-vod .item > a")
            film_links = []
            filmDatas = {}
            for i in range(int(index),len(films_item)) :
                film_links.append(films_item.__getitem__(int(i)).get_attribute("href")) 
            for i in range(int(index),len(film_links)) :
                driver.get(film_links[i])
                try : 
                    filmDatas = self.getFilmsData()
                except : 
                    z=1
           
                driver.execute_script("document.querySelector('#btnPlay').click()")
                time.sleep(5)
                driver.execute_script("document.querySelector('#video').click()")
                driver.find_element_by_css_selector("#video").send_keys(Keys.SPACE)
                
                while(True) : 
                    if(input("Wait for click idm panel to download... Press 1 to confirm pressed") == str(1)) : 
                        break;

                #filmDatas['videoname'] = self.getM3U8File(filmDatas,film_links[i])

                self.pushIntoExel(filmDatas)
                print("Page "+str(page)+" , Index" + str(i))
            driver.get("https://247phim.com/phim/phim-le/nam/2021/trang-{}".format(page))        

    def getM3U8File(self,filmDatas,currenturl) :
        global client
        global server
        global driver
        print(client.proxy)
        count = 0
        global m3u8FileParReg
        while(True) :      
            try : 
                client.new_har()
                time.sleep(1)
                result = json.dumps(client.har)
                data = json.loads(result)
                for entry in data['log']['entries'] : 

                    if(m3u8FileParReg.search(str(entry['request']['url']))) :
                        videoname = unidecode(filmDatas['filmname_vi'].replace(" ",""))
                        return videoname
                            
            except : 
                x = 1
    def getFilmsData(self) :
        global driver
        filmDatas = {}
        time.sleep(5)
        pageSoup = BeautifulSoup(driver.page_source, "html.parser")
        filmDetailSelector = pageSoup.select_one(".detail-vod-left .detail")
        filmDatas['filmtype'] = filmDetailSelector.select_one(".detail-info a:nth-child(2)").find(text=True,recursive=False)
        filmDatas['publishYear'] = filmDetailSelector.select_one(".detail-info a:last-child").find(text=True,recursive=False).replace("'","")
        filmDatas['filmname_vi'] = filmDetailSelector.select_one("h2.title-vod").find(text=True,recursive=False).replace("'","")
        

        filmDatas['filmname_eng'] = filmDetailSelector.select_one("h3.title-vod").find(text=True,recursive=False).replace("'","")
        filmDatas['filmdescription'] =filmDetailSelector.select_one("div.mt-2:nth-child(6) p").find(text=True,recursive=False).replace("'","").replace(u'\xa0', u' ')

        filmDatas['duration'] = [text for text in filmDetailSelector.select_one("div.mt-4:nth-child(7) > div:first-child > ul.more-info > li:first-child").stripped_strings][1]
        
        filmDatas['director'] = filmDetailSelector.select_one("div.mt-4:nth-child(7) > div:first-child ul.more-info li:nth-child(2)").find(text=True,recursive=False).replace("'","")
        filmDatas['country'] = filmDetailSelector.select_one("div.mt-4:nth-child(7) > div:first-child ul.more-info li:nth-child(3)").find(text=True,recursive=False).replace("'","")
        filmDatas['category'] = " ".join(filmDetailSelector.select_one("div.mt-4:nth-child(7) > div:first-child ul.more-info li:nth-child(4)").find(text=True,recursive=False).strip().replace("\n"," ").rstrip("'").replace("'","").replace("\r","").split())
        filmDatas['IMDB'] = filmDetailSelector.select_one("div.mt-4:nth-child(7) > div:first-child ul.more-info li:nth-child(6)").find(text=True,recursive=False).replace("'","")
        filmDatas['cast'] = filmDetailSelector.select_one("div.mt-4:nth-child(7) > div:last-child ul.more-info li").find(text=True,recursive=False).replace("'","")
        return filmDatas
        
    def pushIntoExel(self, filmDatas) : 
        wb = openpyxl.load_workbook("./FilmExel/247PhimLe.xlsx")
        sheet1 = wb.active
        for rowindex in range(sheet1.max_row+1,sheet1.max_row+2) : 
            for columnindex in range(1,sheet1.max_column+1) : 
                
                sheet1.cell(rowindex,column=columnindex).value = filmDatas[sheet1.cell(1,columnindex).value]

        wb.save("./FilmExel/247PhimLe.xlsx")
        wb.close()



crawler = PhimmoiCrawler("https://247phim.com/phim/phim-le/nam/2021/")
crawler.interactSiteToAjaxCome(6)
crawler.GoIntoFilm()

#index 25; page 1